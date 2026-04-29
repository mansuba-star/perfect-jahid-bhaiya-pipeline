import re
import json
import time
import argparse
import sys
import os
from pathlib import Path

_anthropic_available = False
_openai_available    = False
_genai_available     = False

try:
    import anthropic as _anthropic_module
    _anthropic_available = True
except ImportError:
    pass

try:
    import openai as _openai_module
    _openai_available = True
except ImportError:
    pass

try:
    import google.generativeai as genai
    _genai_available = True
except ImportError:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Missing dependency. Run: pip install openpyxl")
    sys.exit(1)
_anthropic = _anthropic_module if _anthropic_available else None

def parse_questions(txt_path: str) -> list[dict]:
    
    text = Path(txt_path).read_text(encoding="utf-8")

    text = text.replace("\r\n", "\n").replace("\r", "\n")

    blocks = re.split(r"\n(?=question\s*:)", text.strip(), flags=re.IGNORECASE)

    questions = []
    for idx, block in enumerate(blocks, start=1):
        block = block.strip()
        if not block:
            continue

        q: dict = {"id": idx, "raw": block}

        field_pattern = re.compile(
            r"^(question|option1|option2|option3|option4|answer|explain|slug)\s*:\s*",
            re.IGNORECASE | re.MULTILINE,
        )

        positions = [(m.group(1).lower(), m.start(), m.end()) for m in field_pattern.finditer(block)]

        for i, (field, _, val_start) in enumerate(positions):
            val_end = positions[i + 1][1] if i + 1 < len(positions) else len(block)
            q[field] = block[val_start:val_end].strip()

        for f in ("question", "option1", "option2", "option3", "option4", "answer", "explain", "slug"):
            q.setdefault(f, "")

        questions.append(q)

    return questions

SUSPICIOUS_YEAR_PAIRS = [
    (r"\b1861\b", "1681"), 
    (r"\b1681\b", "1861"),
    (r"\b1971\b", "1791"),
    (r"\b1947\b", "1974"),
    (r"\b1952\b", "1925"),
    (r"\b1905\b", "1950"),
    (r"\b1206\b", "1260"),
    (r"\b1757\b", "1577"),
    (r"\b1857\b", "1587"),
]

def extract_all_numbers(text: str) -> list[str]:
    return re.findall(r"\b\d+\b", text)


def regex_number_check(q: dict) -> list[dict]:
    """Flag numbers that look like common transpositions."""
    errors = []
    fields_to_check = {
        "question": q.get("question", ""),
        "explain":  q.get("explain", ""),
        "option1":  q.get("option1", ""),
        "option2":  q.get("option2", ""),
        "option3":  q.get("option3", ""),
        "option4":  q.get("option4", ""),
    }

    for field, text in fields_to_check.items():
        for pattern, common_mistake in SUSPICIOUS_YEAR_PAIRS:
            if re.search(pattern, text):
                # Check if the "common mistake" version also appears — if so, one of them is wrong
                if re.search(r"\b" + common_mistake + r"\b", text):
                    errors.append({
                        "type": "NUMBER_CONFLICT",
                        "field": field,
                        "description": f"Both {pattern[2:-2]} and {common_mistake} appear in the same text — one is likely wrong",
                        "wrong_text": f"Conflicting: {pattern[2:-2]} vs {common_mistake}",
                        "suggested_fix": "Verify against authoritative source",
                    })

    return errors


def cross_field_consistency_check(q: dict) -> list[dict]:
    
    errors = []
    answer_text = q.get("answer", "").strip()
    explain_text = q.get("explain", "")

    answer_numbers = re.findall(r"\b\d{3,}\b", answer_text)
    option_numbers = re.findall(r"\b\d{3,}\b", " ".join([
        q.get("option1", ""), q.get("option2", ""),
        q.get("option3", ""), q.get("option4", ""),
    ]))
    all_question_numbers = list(set(answer_numbers + option_numbers))

    explain_numbers = re.findall(r"\b\d{3,}\b", explain_text)

    for qnum in all_question_numbers:
        if not explain_numbers:
            continue
        for enum in explain_numbers:
            
            if (len(qnum) == len(enum) and qnum != enum and
                    sorted(qnum) == sorted(enum)):
                errors.append({
                    "type": "NUMBER_ERROR",
                    "field": "explain",
                    "description": (
                        f"Number in explanation ({enum}) looks like a transposition of "
                        f"a number in the question/answer ({qnum}). "
                        f"One of them is likely wrong."
                    ),
                    "wrong_text": f"explain has '{enum}', question/answer has '{qnum}'",
                    "suggested_fix": f"Check whether the explanation should say '{qnum}'",
                })

    # ── Answer vs options check ────────────────────────────────────────
    options_combined = " ".join([
        q.get("option1", ""), q.get("option2", ""),
        q.get("option3", ""), q.get("option4", ""),
    ]).lower()

    clean_answer = re.sub(r"^[কখগঘabcdABCD\)\s]+", "", answer_text).strip().lower()

    if clean_answer and clean_answer not in options_combined:
        errors.append({
            "type": "CONSISTENCY_ERROR",
            "field": "answer",
            "description": "The declared answer text doesn't seem to match any option text",
            "wrong_text": answer_text,
            "suggested_fix": "Verify the answer matches one of the four options",
        })

    return errors

def find_reference_section(reference_text: str, q: dict,
                            max_chars: int = 6000,
                            top_n_passages: int = 5) -> str | None:

    if not reference_text:
        return None, "", []

    _cluster_re = re.compile(r"Chapter\[?\w*\]?\s*/\s*Cluster\s*\d+.*", re.IGNORECASE)

    all_fields = " ".join([
        q.get("question", ""), q.get("explain",  ""),
        q.get("option1",  ""), q.get("option2",  ""),
        q.get("option3",  ""), q.get("option4",  ""),
        q.get("answer",   ""),
    ])

    word_terms = list(set(w for w in re.findall(r"\S{3,}", all_fields)))[:30]

    _bn_digit_map = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")
    def _normalise_nums(text: str) -> list[str]:
        return re.findall(r"\d+", text.translate(_bn_digit_map))

    question_nums = set(_normalise_nums(all_fields))

    lines = reference_text.split("\n")

    # Track which cluster/chapter header each line belongs to
    line_cluster: list[str] = []
    current_cluster = ""
    for line in lines:
        if _cluster_re.search(line):
            current_cluster = line.strip()
        line_cluster.append(current_cluster)

    # Score every individual line
    line_to_score: dict[int, int] = {}
    scores: list[tuple[float, int]] = []

    for i, line in enumerate(lines):
        if not line.strip():
            continue
        word_score = sum(1 for t in word_terms if t in line)
        line_nums  = set(_normalise_nums(line))
        num_score  = sum(3 for n in question_nums if n in line_nums)
        total = word_score + num_score
        if total > 0:
            line_to_score[i] = total
            scores.append((total, i))

    if not scores:
        return None, "", []

    scores.sort(key=lambda x: x[0], reverse=True)

    CONTEXT_BEFORE  = 5
    CONTEXT_AFTER   = 18
    used_lines: set[int] = set()
    passages: list[str] = []
    scored_passages: list[dict] = []
    primary_cluster = ""

    for best_score, best_line in scores:
        if best_line in used_lines:
            continue
        start   = max(0, best_line - CONTEXT_BEFORE)
        end     = min(len(lines), best_line + CONTEXT_AFTER)
        cluster = line_cluster[best_line]
        if not primary_cluster and cluster:
            primary_cluster = cluster

        # Build per-sentence scored lines for this passage
        scored_lines: list[dict] = []
        for li in range(start, end):
            txt = lines[li].strip()
            if not txt:
                continue
            scored_lines.append({
                "score": line_to_score.get(li, 0),
                "text":  txt,
            })

        passage = "\n".join(lines[start:end]).strip()
        if passage:
            header = f"[{cluster}]\n" if cluster else ""
            passages.append(f"{header}{passage}")
            scored_passages.append({
                "score":        best_score,
                "cluster":      cluster,
                "scored_lines": scored_lines,
            })
            used_lines.update(range(start, end))
        if len(passages) >= top_n_passages:
            break

    combined = "\n\n--- (next relevant section) ---\n\n".join(passages)
    return combined[:max_chars], primary_cluster, scored_passages

AI_PROMPT_TEMPLATE = """\
You are an expert fact-checker for Bengali educational Q&A content (MCQ format).

Carefully analyze the question block below and detect any errors.

--- QUESTION BLOCK ---
Question  : {question}
Option 1  : {option1}
Option 2  : {option2}
Option 3  : {option3}
Option 4  : {option4}
Answer    : {answer}
Explanation: {explain}
--- END BLOCK ---
{reference_section}

IMPORTANT INSTRUCTIONS:
- If a REFERENCE SOURCE is provided above, treat every fact in it as ground truth.
  Any number, name, or count in the question block that differs from the reference
  MUST be flagged as an error — even if the difference seems small (e.g. ২৩ vs ১৩,
  তিন vs চার).
- Pay special attention to Bengali and ASCII numerals (০-৯ / 0-9): they are the
  most common source of transposition errors.
- Do NOT rely on your own training knowledge when a reference is supplied. If the
  reference says "২৩টি ভাষায়" and the question says "১৩টি ভাষায়", that is a
  FACTUAL_ERROR regardless of what you may believe is correct.
- Check ALL fields: question, every option, answer, and explanation.

OUTPUT RULES — follow exactly:
- "wrong_text": Copy the EXACT Bengali sentence or phrase from the question block
  that contains the error. Do NOT paraphrase. Quote it verbatim as it appears.
- "description": State clearly what is wrong — quote the wrong part and the correct
  part side by side. Example: 'ব্যাখ্যায় লেখা "রামনিধি গুপ্ত" কবিগানের রচয়িতা হিসেবে
  উল্লেখ করা হয়েছে, কিন্তু তিনি টপ্পা সংগীতের পথিকৃৎ।'
- "reference_evidence": Copy the EXACT line or sentence from the REFERENCE SOURCE
  that proves the error. If no reference is provided leave empty string.
- "suggested_fix": The corrected Bengali text that should replace wrong_text.

Check for ALL of the following error types:

1. NUMBER_ERROR  — Transposed/wrong digits in years, counts, page numbers
   Example: "১৩" written instead of "২৩", "1681" instead of "1861"

2. SPELLING_ERROR — Misspelled Bengali or English words
   Example: "রবীন্দ্রনাথ" misspelled, wrong transliteration

3. FACTUAL_ERROR — A stated fact contradicts the reference source or well-known history
   Example: wrong birth year, wrong district, wrong dynasty, wrong number of poems,
            wrong number of languages, wrong number of types

4. CONSISTENCY_ERROR — Answer doesn't match options, or explanation contradicts the answer


Respond ONLY with valid JSON, exactly in this schema (no extra text):
{{
  "has_errors": true,
  "errors": [
    {{
      "type": "NUMBER_ERROR | SPELLING_ERROR | FACTUAL_ERROR | CONSISTENCY_ERROR",
      "field": "question | option1 | option2 | option3 | option4 | answer | explain",
      "description": "Exact Bengali description — quote wrong part and correct part side by side",
      "wrong_text": "Exact verbatim Bengali sentence or phrase from the question block",
      "reference_cluster": "The cluster/chapter header label e.g. 'Chapter[a] / Cluster 01 — Ode' from which the evidence was taken",
      "reference_evidence": "Exact verbatim line from the REFERENCE SOURCE that proves the error",
      "suggested_fix": "Corrected Bengali text that should replace wrong_text"
    }}
  ],
  "confidence": "HIGH | MEDIUM | LOW",
  "notes": "Any additional observations"
}}

If there are NO errors, return:
{{"has_errors": false, "errors": [], "confidence": "HIGH", "notes": ""}}
"""

def _build_prompt(q: dict, reference_section: str | None) -> str:
    """Build the shared AI prompt string."""
    ref_part = ""
    if reference_section:
        # Count how many passages were returned so the LLM knows scope
        passage_count = reference_section.count("--- (next relevant section) ---") + 1
        ref_part = (
            f"\n--- REFERENCE SOURCE ({passage_count} passage(s) — treat as ground truth) ---\n"
            f"{reference_section}\n"
            f"--- END REFERENCE ---\n"
        )
    return AI_PROMPT_TEMPLATE.format(
        question          = q.get("question", ""),
        option1           = q.get("option1", ""),
        option2           = q.get("option2", ""),
        option3           = q.get("option3", ""),
        option4           = q.get("option4", ""),
        answer            = q.get("answer", ""),
        explain           = q.get("explain", ""),
        reference_section = ref_part,
    )


def _parse_ai_json(raw: str, provider: str) -> dict:
    """Robustly extract and parse the JSON block from an AI response."""
    try:
        json_match = re.search(r"\{[\s\S]*\}", raw)
        if json_match:
            return json.loads(json_match.group())
    except json.JSONDecodeError as e:
        return {"has_errors": False, "errors": [], "confidence": "LOW",
                "notes": f"{provider} JSON parse error: {e}"}
    return {"has_errors": False, "errors": [], "confidence": "LOW",
            "notes": f"Could not parse {provider} response"}

def check_with_claude(q: dict, reference_section: str | None, client, model: str) -> dict:
    prompt = _build_prompt(q, reference_section)
    try:
        message = client.messages.create(
            model      = model,
            max_tokens = 1024,
            messages   = [{"role": "user", "content": prompt}],
        )
        result = _parse_ai_json(message.content[0].text.strip(), "Claude")
        result["_usage"] = {
            "input_tokens":  getattr(message.usage, "input_tokens",  0),
            "output_tokens": getattr(message.usage, "output_tokens", 0),
        }
        return result
    except Exception as e:
        return {"has_errors": False, "errors": [], "confidence": "LOW",
                "notes": f"Claude error: {e}", "_usage": {"input_tokens": 0, "output_tokens": 0}}

def check_with_openai(q: dict, reference_section: str | None, client, model: str) -> dict:
    prompt = _build_prompt(q, reference_section)
    try:
        response = client.chat.completions.create(
            model      = model,
            max_tokens = 1024,
            messages   = [{"role": "user", "content": prompt}],
        )
        raw = response.choices[0].message.content.strip()
        result = _parse_ai_json(raw, "OpenAI")
        usage  = response.usage
        result["_usage"] = {
            "input_tokens":  getattr(usage, "prompt_tokens",     0),
            "output_tokens": getattr(usage, "completion_tokens", 0),
        }
        return result
    except Exception as e:
        return {"has_errors": False, "errors": [], "confidence": "LOW",
                "notes": f"OpenAI error: {e}", "_usage": {"input_tokens": 0, "output_tokens": 0}}


def check_with_gemini(q: dict, reference_section: str | None, model_name: str) -> dict:
    """Gemini uses a module-level configure; no client object is passed."""
    prompt = _build_prompt(q, reference_section)
    try:
        model    = genai.GenerativeModel(model_name)
        response = model.generate_content(prompt)
        raw      = response.text.strip()
        result   = _parse_ai_json(raw, "Gemini")
        meta     = getattr(response, "usage_metadata", None)
        result["_usage"] = {
            "input_tokens":  getattr(meta, "prompt_token_count",     0) if meta else 0,
            "output_tokens": getattr(meta, "candidates_token_count", 0) if meta else 0,
        }
        return result
    except Exception as e:
        return {"has_errors": False, "errors": [], "confidence": "LOW",
                "notes": f"Gemini error: {e}", "_usage": {"input_tokens": 0, "output_tokens": 0}}

def check_with_grok(q: dict, reference_section: str | None, client, model: str) -> dict:
    """Grok uses the same OpenAI client pointed at xAI's base URL."""
    prompt = _build_prompt(q, reference_section)
    try:
        response = client.chat.completions.create(
            model      = model,
            max_tokens = 1024,
            messages   = [{"role": "user", "content": prompt}],
        )
        raw    = response.choices[0].message.content.strip()
        result = _parse_ai_json(raw, "Grok")
        usage  = response.usage
        result["_usage"] = {
            "input_tokens":  getattr(usage, "prompt_tokens",     0),
            "output_tokens": getattr(usage, "completion_tokens", 0),
        }
        return result
    except Exception as e:
        return {"has_errors": False, "errors": [], "confidence": "LOW",
                "notes": f"Grok error: {e}", "_usage": {"input_tokens": 0, "output_tokens": 0}}

def merge_results(regex_errors: list, consistency_errors: list, ai_result: dict) -> dict:
    """Merge all error sources into a single result dict."""
    all_errors = list(regex_errors) + list(consistency_errors)

    if ai_result.get("has_errors"):
        all_errors.extend(ai_result.get("errors", []))

    seen = set()
    unique_errors = []
    for e in all_errors:
        key = (e.get("type"), e.get("field"), e.get("wrong_text", "")[:60])
        if key not in seen:
            seen.add(key)
            unique_errors.append(e)

    return {
        "has_errors": bool(unique_errors),
        "errors": unique_errors,
        "confidence": ai_result.get("confidence", "LOW"),
        "notes": ai_result.get("notes", ""),
    }

C_HEADER_DARK  = "1F4E79"   # dark blue — header rows
C_HEADER_MID   = "2E75B6"   # mid blue — sub-headers
C_ERROR_RED    = "FFD7D7"   # light red — rows with errors
C_OK_GREEN     = "E2EFDA"   # light green — clean rows
C_WARNING_YLW  = "FFF2CC"   # yellow — medium confidence
C_NUMBER_ERR   = "FF6B6B"   # red chip — number errors
C_SPELL_ERR    = "FFD93D"   # yellow chip — spelling
C_FACT_ERR     = "FF8C42"   # orange chip — factual
C_CONSIST_ERR  = "C084FC"   # purple chip — consistency
C_CONFLICT_ERR = "FF6B6B"   # red — number conflict

ERROR_COLOR_MAP = {
    "NUMBER_ERROR":     C_NUMBER_ERR,
    "NUMBER_CONFLICT":  C_CONFLICT_ERR,
    "SPELLING_ERROR":   C_SPELL_ERR,
    "FACTUAL_ERROR":    C_FACT_ERR,
    "CONSISTENCY_ERROR": C_CONSIST_ERR,
}

def hdr_cell(ws, row, col, value, bg=C_HEADER_DARK, fg="FFFFFF", size=11, bold=True, wrap=True, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell


def body_cell(ws, row, col, value, bg=None, bold=False, wrap=True, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    return cell


def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_excel_report(questions: list[dict], results: list[dict], output_path: str,
                           usage_summary: dict | None = None):
    wb = openpyxl.Workbook()

    total        = len(questions)
    with_errors  = sum(1 for r in results if r and r.get("has_errors"))
    error_rate   = (with_errors / total * 100) if total else 0

    # ── Sheet 1: Dashboard ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "Question Error Detection Report"
    ws["A1"].font      = Font(name="Arial", bold=True, size=18, color=C_HEADER_DARK)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Analysed {total} questions  •  {with_errors} flagged  •  Error rate {error_rate:.1f}%"
    ws["A2"].font      = Font(name="Arial", size=11, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    stat_rows = [
        ("Total Questions",         total,                  C_HEADER_MID, "FFFFFF"),
        ("Questions with Errors",   with_errors,            "C00000",     "FFFFFF"),
        ("Clean Questions",         total - with_errors,    "375623",     "FFFFFF"),
        ("Error Rate",              f"{error_rate:.1f}%",   "833C00",     "FFFFFF"),
    ]
    for i, (label, value, bg, fg) in enumerate(stat_rows, start=4):
        ws.merge_cells(f"B{i}:C{i}")
        ws.merge_cells(f"D{i}:E{i}")
        lc = ws.cell(row=i, column=2, value=label)
        lc.font      = Font(name="Arial", bold=True, size=11, color=fg)
        lc.fill      = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc = ws.cell(row=i, column=4, value=value)
        vc.font      = Font(name="Arial", bold=True, size=13, color=fg)
        vc.fill      = PatternFill("solid", start_color=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 24

    ws["B11"] = "Error Type Breakdown"
    ws["B11"].font = Font(name="Arial", bold=True, size=12, color=C_HEADER_DARK)

    type_counts: dict[str, int] = {}
    for r in results:
        if r:
            for e in r.get("errors", []):
                t = e.get("type", "UNKNOWN")
                type_counts[t] = type_counts.get(t, 0) + 1

    hdr_cell(ws, 12, 2, "Error Type",  bg=C_HEADER_MID)
    hdr_cell(ws, 12, 3, "Count",       bg=C_HEADER_MID)
    hdr_cell(ws, 12, 4, "% of Total",  bg=C_HEADER_MID)

    for ri, (etype, count) in enumerate(sorted(type_counts.items()), start=13):
        pct = count / total * 100 if total else 0
        bg  = ERROR_COLOR_MAP.get(etype, "CCCCCC")
        body_cell(ws, ri, 2, etype,         bg=bg, bold=True)
        body_cell(ws, ri, 3, count,         bg=bg, align="center")
        body_cell(ws, ri, 4, f"{pct:.1f}%", bg=bg, align="center")

    # ── Token Usage & Cost section ──────────────────────────────────────
    if usage_summary:
        next_row = 13 + len(type_counts) + 2

        ws.cell(row=next_row, column=2, value="Token Usage & Cost").font = Font(
            name="Arial", bold=True, size=12, color=C_HEADER_DARK)
        next_row += 1

        hdr_cell(ws, next_row, 2, "Metric",  bg="4472C4")
        hdr_cell(ws, next_row, 3, "Value",   bg="4472C4")
        next_row += 1

        cost_rows = [
            ("Provider",          usage_summary.get("provider", "").upper()),
            ("Model",             usage_summary.get("model", "")),
            ("Input tokens",      f"{usage_summary.get('input_tokens', 0):,}"),
            ("Output tokens",     f"{usage_summary.get('output_tokens', 0):,}"),
            ("Total tokens",      f"{usage_summary.get('input_tokens', 0) + usage_summary.get('output_tokens', 0):,}"),
            ("Input cost",        f"${usage_summary.get('input_cost_usd', 0):.6f}"),
            ("Output cost",       f"${usage_summary.get('output_cost_usd', 0):.6f}"),
            ("Total cost (USD)",  f"${usage_summary.get('total_cost_usd', 0):.6f}"),
        ]
        for label, val in cost_rows:
            body_cell(ws, next_row, 2, label, bg="DDEEFF", bold=True)
            body_cell(ws, next_row, 3, val,   bg="DDEEFF", align="right")
            ws.row_dimensions[next_row].height = 20
            next_row += 1

        if usage_summary.get("total_cost_usd", 0) == 0.0 and usage_summary.get("input_tokens", 0) > 0:
            ws.cell(row=next_row, column=2,
                    value="* Pricing unavailable for this model — token counts are accurate.").font = Font(
                name="Arial", italic=True, size=9, color="595959")

    set_col_widths(ws, [3, 30, 16, 20, 12, 3])

    # ── Sheet 2: All Questions ──────────────────────────────────────────
    ws2 = wb.create_sheet("📋 All Questions")
    ws2.sheet_view.showGridLines = False

    headers2 = ["#", "Question", "Slug", "Answer", "Status", "Error Count", "Error Types", "Confidence", "Explanation", "Notes"]
    for ci, h in enumerate(headers2, 1):
        hdr_cell(ws2, 1, ci, h)
    ws2.freeze_panes = "A2"

    for ri, (q, result) in enumerate(zip(questions, results), start=2):
        r = result or {}
        has_err    = r.get("has_errors", False)
        errors     = r.get("errors", [])
        err_types  = ", ".join(sorted(set(e.get("type", "") for e in errors))) if errors else "—"
        row_bg     = C_ERROR_RED if has_err else C_OK_GREEN

        body_cell(ws2, ri, 1,  q["id"],                              bg=row_bg, align="center")
        body_cell(ws2, ri, 2,  q["question"][:200],                  bg=row_bg)
        body_cell(ws2, ri, 3,  q.get("slug", ""),                    bg=row_bg)
        body_cell(ws2, ri, 4,  q["answer"],                          bg=row_bg)
        body_cell(ws2, ri, 5,  "⚠ ERROR" if has_err else "✓ OK",    bg=row_bg, bold=has_err, align="center")
        body_cell(ws2, ri, 6,  len(errors),                          bg=row_bg, align="center")
        body_cell(ws2, ri, 7,  err_types,                            bg=row_bg)
        body_cell(ws2, ri, 8,  r.get("confidence", ""),              bg=row_bg, align="center")
        body_cell(ws2, ri, 9,  q.get("explain", ""),                  bg=row_bg)
        body_cell(ws2, ri, 10, r.get("notes", "")[:120],             bg=row_bg)

        ws2.row_dimensions[ri].height = 55

    set_col_widths(ws2, [5, 55, 25, 22, 10, 12, 38, 12, 55, 45])

    # ── Sheet 3: Detailed Errors ────────────────────────────────────────
    ws3 = wb.create_sheet("Detailed Errors")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Q#", "Slug", "Question (excerpt)", "Error Type", "Field", "Wrong Text (Bengali)", "Cluster / Source", "Retrieved Passages (Score)", "Reference Evidence", "Suggested Fix", "Description", "Explanation"]
    for ci, h in enumerate(headers3, 1):
        hdr_cell(ws3, 1, ci, h)
    ws3.freeze_panes = "A2"

    err_row = 2
    for q, result in zip(questions, results):
        if not result or not result.get("has_errors"):
            continue
        fallback_cluster  = result.get("_primary_cluster", "")
        scored_passages   = result.get("_scored_passages", [])

        # Format all 5 retrieved passages with per-sentence scores
        passages_cell = ""
        for idx, sp in enumerate(scored_passages, 1):
            cluster_lbl  = sp.get("cluster", "")
            best_score   = sp.get("score", 0)
            scored_lines = sp.get("scored_lines", [])
            passages_cell += f"{'─'*48}\n"
            passages_cell += f"#{idx}  Passage Score: {best_score}  |  {cluster_lbl}\n"
            passages_cell += f"{'─'*48}\n"
            for sl in scored_lines:
                s   = sl.get("score", 0)
                txt = sl.get("text", "")
                prefix = f"[{s:>2}] " if s > 0 else "     "
                passages_cell += f"{prefix}{txt}\n"
            passages_cell += "\n"

        for error in result.get("errors", []):
            etype   = error.get("type", "UNKNOWN")
            row_bg  = ERROR_COLOR_MAP.get(etype, "DDDDDD")
            excerpt = q["question"][:100] + ("…" if len(q["question"]) > 100 else "")
            cluster = error.get("reference_cluster", "") or fallback_cluster

            body_cell(ws3, err_row, 1,  q["id"],                                bg=row_bg, align="center", bold=True)
            body_cell(ws3, err_row, 2,  q.get("slug", ""),                      bg=row_bg)
            body_cell(ws3, err_row, 3,  excerpt,                                 bg=row_bg)
            body_cell(ws3, err_row, 4,  etype,                                   bg=row_bg, bold=True, align="center")
            body_cell(ws3, err_row, 5,  error.get("field", ""),                  bg=row_bg, align="center")
            body_cell(ws3, err_row, 6,  error.get("wrong_text", ""),             bg=row_bg)
            body_cell(ws3, err_row, 7,  cluster,                                 bg="FFF9E6", bold=True)
            body_cell(ws3, err_row, 8,  passages_cell.strip(),                   bg="EAF4FB")
            body_cell(ws3, err_row, 9,  error.get("reference_evidence", ""),     bg="FFF9E6")
            body_cell(ws3, err_row, 10, error.get("suggested_fix", ""),          bg=row_bg)
            body_cell(ws3, err_row, 11, error.get("description", ""),            bg=row_bg)
            body_cell(ws3, err_row, 12, q.get("explain", ""),                    bg=row_bg)

            ws3.row_dimensions[err_row].height = 250
            err_row += 1

    if err_row == 2:  # No errors found at all
        ws3.cell(row=2, column=1, value="No errors detected ✓").font = Font(
            name="Arial", bold=True, color="375623", size=12)

    set_col_widths(ws3, [5, 25, 45, 20, 12, 45, 30, 60, 55, 40, 60, 60])

    ws4 = wb.create_sheet("Error Question Detail")
    ws4.sheet_view.showGridLines = False

    detail_row = 1
    for q, result in zip(questions, results):
        if not result or not result.get("has_errors"):
            continue

        # Question header
        ws4.merge_cells(f"A{detail_row}:F{detail_row}")
        hdr_cell(ws4, detail_row, 1, f"Q#{q['id']}: {q['question'][:120]}", bg=C_HEADER_DARK, align="left")
        ws4.row_dimensions[detail_row].height = 30
        detail_row += 1

        # Slug row
        if q.get("slug"):
            c1 = ws4.cell(row=detail_row, column=1, value="Slug")
            c1.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_DARK)
            c1.fill      = PatternFill("solid", start_color="FFF2CC")
            c1.alignment = Alignment(horizontal="right", vertical="top", indent=1)
            ws4.merge_cells(f"B{detail_row}:F{detail_row}")
            c2 = ws4.cell(row=detail_row, column=2, value=q["slug"])
            c2.font      = Font(name="Arial", size=10, bold=True)
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws4.row_dimensions[detail_row].height = 22
            detail_row += 1

        # Fields
        field_rows = [
            ("Option 1",    q["option1"]),
            ("Option 2",    q["option2"]),
            ("Option 3",    q["option3"]),
            ("Option 4",    q["option4"]),
            ("Answer",      q["answer"]),
            ("Explanation", q["explain"]),
        ]
        for label, val in field_rows:
            c1 = ws4.cell(row=detail_row, column=1, value=label)
            c1.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_DARK)
            c1.fill      = PatternFill("solid", start_color="DDEEFF")
            c1.alignment = Alignment(horizontal="right", vertical="top", indent=1)

            ws4.merge_cells(f"B{detail_row}:F{detail_row}")
            c2 = ws4.cell(row=detail_row, column=2, value=val)
            c2.font      = Font(name="Arial", size=10)
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws4.row_dimensions[detail_row].height = 35 if label != "Explanation" else 80
            detail_row += 1

        # Errors for this question
        fallback_cluster = result.get("_primary_cluster", "")
        for error in result.get("errors", []):
            etype         = error.get("type", "UNKNOWN")
            row_bg        = ERROR_COLOR_MAP.get(etype, "DDDDDD")
            wrong         = error.get("wrong_text", "")
            fix           = error.get("suggested_fix", "")
            desc          = error.get("description", "")
            ref_ev        = error.get("reference_evidence", "")
            cluster_label = error.get("reference_cluster", "") or fallback_cluster

            ws4.merge_cells(f"A{detail_row}:F{detail_row}")
            summary = f"  ⚠ [{etype}] | ক্ষেত্র: {error.get('field','')} | ভুল: {wrong} → সঠিক: {fix}"
            ec = ws4.cell(row=detail_row, column=1, value=summary)
            ec.font      = Font(name="Arial", bold=True, size=10)
            ec.fill      = PatternFill("solid", start_color=row_bg)
            ec.alignment = Alignment(wrap_text=True, vertical="center")
            ws4.row_dimensions[detail_row].height = 35
            detail_row += 1

            if desc:
                ws4.merge_cells(f"A{detail_row}:F{detail_row}")
                dc = ws4.cell(row=detail_row, column=1, value=f"  বিবরণ: {desc}")
                dc.font      = Font(name="Arial", size=10, italic=True)
                dc.fill      = PatternFill("solid", start_color="F0F0F0")
                dc.alignment = Alignment(wrap_text=True, vertical="top")
                ws4.row_dimensions[detail_row].height = 35
                detail_row += 1

            cluster_part = f"[{cluster_label}]  " if cluster_label else ""
            ws4.merge_cells(f"A{detail_row}:F{detail_row}")
            rc = ws4.cell(row=detail_row, column=1, value=f"  📖 {cluster_part}{ref_ev}")
            rc.font      = Font(name="Arial", size=10, color="7B3F00")
            rc.fill      = PatternFill("solid", start_color="FFF9E6")
            rc.alignment = Alignment(wrap_text=True, vertical="top")
            ws4.row_dimensions[detail_row].height = 40
            detail_row += 1

        # Spacer
        detail_row += 1

    set_col_widths(ws4, [14, 40, 20, 20, 20, 20])

    wb.save(output_path)
    print(f"\n✅ Report saved → {output_path}")

_PRICING: dict[str, tuple[float, float]] = {
    # Claude  (input/M, output/M)
    "claude:claude-opus-4-5":     (15.00,  75.00),
    "claude:claude-opus-4":       (15.00,  75.00),
    "claude:claude-sonnet-4-5":   ( 3.00,  15.00),
    "claude:claude-sonnet-4":     ( 3.00,  15.00),
    "claude:claude-haiku-3-5":    ( 0.80,   4.00),
    "claude:claude-haiku-3":      ( 0.25,   1.25),
    # OpenAI
    "openai:gpt-4o":              ( 2.50,  10.00),
    "openai:gpt-4-turbo":         (10.00,  30.00),
    "openai:gpt-4o-mini":         ( 0.15,   0.60),
    "openai:gpt-3.5-turbo":       ( 0.50,   1.50),
    # Gemini
    "gemini:gemini-2.5-pro":      ( 1.25,  10.00),
    "gemini:gemini-1.5-pro":      ( 1.25,   5.00),
    "gemini:gemini-2.0-flash":    ( 0.10,   0.40),
    "gemini:gemini-1.5-flash":    ( 0.075,  0.30),
    # Grok (xAI)
    "grok:grok-3":                ( 3.00,  15.00),
    "grok:grok-3-mini":           ( 0.30,   0.50),
    "grok:grok-2":                ( 2.00,  10.00),
}


def _get_price_per_token(provider: str, model: str) -> tuple[float, float]:
    """Return (input_$/token, output_$/token). Returns (0, 0) for unknown models."""
    key = f"{provider}:{model}"
    # Exact match
    if key in _PRICING:
        inp_m, out_m = _PRICING[key]
        return inp_m / 1_000_000, out_m / 1_000_000
    # Prefix match — handles versioned names like "gemini-2.5-pro-preview-05-06"
    for k, (inp_m, out_m) in _PRICING.items():
        if key.startswith(k):
            return inp_m / 1_000_000, out_m / 1_000_000
    return 0.0, 0.0


# Default models per provider
_DEFAULT_MODELS = {
    "claude": "claude-sonnet-4-5",
    "openai": "gpt-4o",
    "gemini": "gemini-2.5-pro",
    "grok":   "grok-3-mini",
}

# Environment variable names per provider
_ENV_KEYS = {
    "claude": "ANTHROPIC_API_KEY",
    "openai": "OPENAI_API_KEY",
    "gemini": "GEMINI_API_KEY",
    "grok":   "GROK_API_KEY",
}


def main():
    parser = argparse.ArgumentParser(
        description="Detect errors in Bengali MCQ files using AI (Claude/OpenAI/Gemini/Grok) + regex",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--questions",  required=True,  help="Path to questions .txt file")
    parser.add_argument("--reference",                  help="Path to trusted reference/summary .txt or .md file")
    parser.add_argument("--output",     default="error_report.xlsx", help="Output Excel file (default: error_report.xlsx)")
    parser.add_argument("--provider",   default="claude",
                        choices=["claude", "openai", "gemini", "grok"],
                        help="AI provider to use (default: claude)")
    parser.add_argument("--model",                      help="Override the default model for the chosen provider")
    parser.add_argument("--api-key",                    help="API key (overrides the provider's env-var)")
    parser.add_argument("--limit",      type=int,       help="Process only first N questions (useful for testing)")
    parser.add_argument("--delay",      type=float, default=0.5, help="Seconds to wait between API calls (default: 0.5)")
    parser.add_argument("--skip-ai",    action="store_true",     help="Skip AI checks (regex + consistency only)")
    args = parser.parse_args()

    provider   = args.provider
    model_name = args.model or _DEFAULT_MODELS[provider]
    
    if not Path(args.questions).exists():
        print(f"ERROR: Questions file not found: {args.questions}")
        sys.exit(1)

    reference_text = ""
    if args.reference:
        if not Path(args.reference).exists():
            print(f"WARNING: Reference file not found: {args.reference}  (continuing without it)")
        else:
            reference_text = Path(args.reference).read_text(encoding="utf-8")
            print(f"Loaded reference file: {args.reference}  ({len(reference_text):,} chars)")

    ai_client = None   
    if not args.skip_ai:
        api_key = args.api_key or os.environ.get(_ENV_KEYS[provider])
        if not api_key:
            print(f"WARNING: No API key for provider '{provider}'.")
            print(f"         Set {_ENV_KEYS[provider]} or pass --api-key  or  --skip-ai")
            args.skip_ai = True
        else:
            if provider == "claude":
                if not _anthropic_available:
                    print("ERROR: 'anthropic' package not installed.  Run: pip install anthropic")
                    sys.exit(1)
                ai_client = _anthropic_module.Anthropic(api_key=api_key)

            elif provider == "openai":
                if not _openai_available:
                    print("ERROR: 'openai' package not installed.  Run: pip install openai")
                    sys.exit(1)
                ai_client = _openai_module.OpenAI(api_key=api_key)

            elif provider == "gemini":
                if not _genai_available:
                    print("ERROR: 'google-generativeai' package not installed.  Run: pip install google-generativeai")
                    sys.exit(1)
                genai.configure(api_key=api_key)
                ai_client = True   # genai uses module-level config; just mark as ready

            elif provider == "grok":
                if not _openai_available:
                    print("ERROR: 'openai' package not installed.  Run: pip install openai")
                    sys.exit(1)
                ai_client = _openai_module.OpenAI(
                    api_key  = api_key,
                    base_url = "https://api.x.ai/v1",
                )

    if not args.skip_ai:
        print(f"Provider : {provider.upper()}  |  Model: {model_name}")

    print(f"\nParsing: {args.questions}")
    questions = parse_questions(args.questions)
    if not questions:
        print("ERROR: No questions found. Check that the file uses the expected format.")
        sys.exit(1)

    if args.limit:
        questions = questions[: args.limit]
        print(f"Limiting to first {args.limit} questions (--limit flag)")

    print(f"Parsed {len(questions)} questions\n")

    results: list[dict] = []
    error_count        = 0
    total_input_tokens = 0
    total_output_tokens = 0

    for i, q in enumerate(questions, 1):
        prefix = f"[{i:>4}/{len(questions)}] Q#{q['id']}"

        # Layer 1: Regex number checks
        regex_errors = regex_number_check(q)

        # Layer 2: Cross-field consistency
        consistency_errors = cross_field_consistency_check(q)

        # Layer 3: AI fact-check
        ai_result: dict = {"has_errors": False, "errors": [], "confidence": "HIGH", "notes": ""}
        primary_cluster  = ""
        scored_passages  = []
        if not args.skip_ai and ai_client:
            ref_section, primary_cluster, scored_passages = find_reference_section(reference_text, q) if reference_text else (None, "", [])
            if provider == "claude":
                ai_result = check_with_claude(q, ref_section, ai_client, model_name)
            elif provider == "openai":
                ai_result = check_with_openai(q, ref_section, ai_client, model_name)
            elif provider == "gemini":
                ai_result = check_with_gemini(q, ref_section, model_name)
            elif provider == "grok":
                ai_result = check_with_grok(q, ref_section, ai_client, model_name)
            time.sleep(args.delay)

        usage = ai_result.pop("_usage", {"input_tokens": 0, "output_tokens": 0})
        total_input_tokens  += usage.get("input_tokens",  0)
        total_output_tokens += usage.get("output_tokens", 0)

        # Merge all
        merged = merge_results(regex_errors, consistency_errors, ai_result)
        merged["_primary_cluster"] = primary_cluster
        merged["_scored_passages"] = scored_passages
        results.append(merged)

        if merged["has_errors"]:
            error_count += 1
            types = ", ".join(set(e.get("type", "") for e in merged["errors"]))
            print(f"{prefix}   {len(merged['errors'])} error(s) — {types}")
        else:
            print(f"{prefix} ")

    inp_rate, out_rate = _get_price_per_token(provider, model_name)
    input_cost  = total_input_tokens  * inp_rate
    output_cost = total_output_tokens * out_rate
    total_cost  = input_cost + output_cost
    pricing_known = (inp_rate > 0)

    usage_summary = {
        "provider":       provider,
        "model":          model_name,
        "input_tokens":   total_input_tokens,
        "output_tokens":  total_output_tokens,
        "input_cost_usd": input_cost,
        "output_cost_usd":output_cost,
        "total_cost_usd": total_cost,
    } if (not args.skip_ai and (total_input_tokens or total_output_tokens)) else None

    print(f"\nGenerating Excel report → {args.output}")
    generate_excel_report(questions, results, args.output, usage_summary=usage_summary)

    total_tokens = total_input_tokens + total_output_tokens
    print(f"\n{'─'*55}")
    print(f"  Total questions checked : {len(questions)}")
    print(f"  Questions with errors   : {error_count}")
    print(f"  Clean questions         : {len(questions) - error_count}")
    print(f"  Error rate              : {error_count/len(questions)*100:.1f}%")
    if usage_summary:
        print(f"{'─'*55}")
        print(f"  Provider / Model        : {provider.upper()} / {model_name}")
        print(f"  Input  tokens           : {total_input_tokens:,}")
        print(f"  Output tokens           : {total_output_tokens:,}")
        print(f"  Total  tokens           : {total_tokens:,}")
        if pricing_known:
            print(f"  Input  cost             : ${input_cost:.6f}")
            print(f"  Output cost             : ${output_cost:.6f}")
            print(f"  ► Total cost (USD)      : ${total_cost:.6f}")
        else:
            print(f"  Cost                    : (pricing unknown for '{model_name}')")
    print(f"  Report                  : {args.output}")
    print(f"{'─'*55}\n")


if __name__ == "__main__":
    main()